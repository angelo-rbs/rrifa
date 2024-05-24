import random

import openpyxl

# contagem das colunas começa de 1
NUM_COLUNA_IDENTIFICADOR = 2
NUM_COLUNA_PESO = 4

# presume que estão na mesma pasta do script
NOMES_TABELAS = []
COLUNAS_TEM_CABECALHO = True

fichas = []


for caminho in NOMES_TABELAS:
	projeto = openpyxl.load_workbook(caminho)
	planilha = projeto.active

	linhas = planilha.max_row
	colunas = planilha.max_column

	
	linha_inicial = 2 if COLUNAS_TEM_CABECALHO else 1
    
	for i in range(linha_inicial, linha_inicial + linhas - 1):

		
		nome = planilha.cell(row=i, column=NUM_COLUNA_IDENTIFICADOR).value
		print(nome)
		
		fichas_compradas = int(planilha.cell(row=i, column=NUM_COLUNA_PESO).value)

		for j in range(fichas_compradas):
			fichas.append(nome)


numero = random.randint(0, len(fichas))
vencedor = fichas[numero]


print()
print(len(fichas), "fichas")
print("vencedor sorteado:", vencedor)
